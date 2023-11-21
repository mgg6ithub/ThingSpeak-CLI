import os
import openpyxl
import time
# import pdb


# feeds = [{'created_at': '2023-11-08T22:48:56Z', 'entry_id': 1, 'field1': '0.1', 'field2': None}, 
#         {'created_at': '2023-11-08T22:49:12Z', 'entry_id': 2, 'field1': '0.4', 'field2': None}, 
#         {'created_at': '2023-11-08T22:49:27Z', 'entry_id': 3, 'field1': '0.4', 'field2': None}, 
#         {'created_at': '2023-11-08T22:49:43Z', 'entry_id': 4, 'field1': '0.3', 'field2': None}, 
#         {'created_at': '2023-11-08T22:49:59Z', 'entry_id': 5, 'field1': '0.2', 'field2': None}, 
#         {'created_at': '2023-11-08T22:50:14Z', 'entry_id': 6, 'field1': '0.2', 'field2': None}, 
#         {'created_at': '2023-11-08T22:50:30Z', 'entry_id': 7, 'field1': '0.2', 'field2': None}, 
#         {'created_at': '2023-11-09T16:13:33Z', 'entry_id': 8, 'field1': None, 'field2': '0.1'}, 
#         {'created_at': '2023-11-09T16:13:49Z', 'entry_id': 9, 'field1': None, 'field2': '0.3'}, 
#         {'created_at': '2023-11-09T16:14:04Z', 'entry_id': 10, 'field1': None, 'field2': '0.2'}, 
#         {'created_at': '2023-11-09T16:14:20Z', 'entry_id': 11, 'field1': None, 'field2': '0.3'}, 
#         {'created_at': '2023-11-09T16:14:35Z', 'entry_id': 12, 'field1': None, 'field2': '0.4'}, 
#         {'created_at': '2023-11-09T16:14:51Z', 'entry_id': 13, 'field1': None, 'field2': '0.2'}, 
#         {'created_at': '2023-11-09T16:15:06Z', 'entry_id': 14, 'field1': None, 'field2': '0.2'}]


# def introducir_fila_excel(ws, fila, datos):
#     if type(datos) is list:
#         # pdb.set_trace()
#         col = 1
#         for d in datos:
#             ws.cell(row=fila, column=col, value=d)
#             col += 1
#     else:
#         ws.cell(row=fila, column=1, value=datos)

# file_name = "test"

# store_path = os.getcwd() + "/" + file_name + ".xlsx"

# try:
#     wb = openpyxl.load_workbook(store_path)
# except FileNotFoundError:
#     wb = openpyxl.Workbook()

# ws = wb.active

# ws.title = file_name

# row = 2
# for data_row in feeds:
#     introducir_fila_excel(ws, row, [data_row['created_at'], data_row['field1']])
#     row += 1

# introducir_fila_excel(ws, 1, "PRUEBA")

# wb.save(file_name + ".xlsx")

# print(f"{file_name}.xlsx created.")
# time.sleep(3)

with open('backup.csv', 'r') as file:
    file_data = file.readlines()
data_values = []
for data in file_data:
    data_values.append(data.split('\t')[2].strip())

print(data_values)