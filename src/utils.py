
import os
import requests
from requests.exceptions import HTTPError, ConnectionError, InvalidSchema, InvalidURL
import platform
import time
import json
from colorama import Fore
from tabulate import tabulate
from datetime import datetime
import pdb
import pandas as pd
import openpyxl
import re

# Define la variable global
clear_command = "cls" if platform.system() == "Windows" else "clear"
menu_stack = []


class Utils:

    def __init__(self):
        self.clear_command = "cls" if platform.system() == "Windows" else "clear"


    # Clear screen method
    @staticmethod
    def clear():
        return_code = os.system(clear_command)
        if return_code != 0:
            print("Error al limpiar la pantalla")


    @staticmethod
    def printRequest(req):
        print(req.status_code)
        print(req.json())


    @staticmethod
    def printFormatedTable(tableHeaders, tableData):
        return tabulate([tableHeaders, *tableData], headers="firstrow", tablefmt="rounded_grid", stralign="center", colalign=("center",))


    @staticmethod
    # Wait method
    def wait(t=None, filename=None):
        try:
            if t is None:
                print(filename + " created.")
                time.sleep(2)
            else:
                time.sleep(t)
        except KeyboardInterrupt:
            print("Has interrumpido la espera del programa.\n")


    @staticmethod
    def hide_cursor():
        print("\x1b[?25l")  # hidden


    @staticmethod
    def show_cursor():
        print("\x1b[?25h")  # shown


    # Wait and hide cursor
    @staticmethod  
    def wait_animation(time_to_wait):
        Utils.hide_cursor()
        Utils.wait(time_to_wait)
        Utils.show_cursor()


    # Endless ThingSpeak-CLI terminal
    @staticmethod
    def endless_terminal(message, *options, help_message=None, menu=None, menu1=None, clear=False, tty=True, exit=False, only_string=False):

        if clear:
            Utils.clear()

        if not tty:
            return input(message)

        print(message)

        while True:
            if menu1:
                i = str(input(Fore.GREEN + f"[{menu}] [{menu1}] ts> " + Fore.WHITE))
            elif menu:
                i = str(input(Fore.GREEN + f"[{menu}] ts> " + Fore.WHITE))
            else:
                i = str(input(Fore.GREEN + "ts> " + Fore.WHITE))
            if i == 'clear' or i == 'cls':
                Utils.clear()
            if i == 'help' and help_message:
                print("\n" + help_message)
            if only_string:
                return i
            if i in options or i.__eq__("b") or exit:
                return i                


    # Metodo para convertir una lista a un objeto json
    @staticmethod
    def list_to_json(lista):
        return json.dumps(lista)
    

    # Method to give the client a status response CORRECT|ERROR
    @staticmethod
    def give_response(message=None, clear=False, status=True):
        
        if clear:
            Utils.clear()

        if message:
            print(message, end='')

        if status:
            print(Fore.GREEN + " successfull." + Fore.WHITE)
            Utils.wait(2)
        else:
            print(Fore.RED + " error." + Fore.WHITE)
            Utils.wait(2)


    # Method to make http requests
    @staticmethod
    def make_request(**kwargs):
        try:
            r = requests.request(**kwargs)
        except InvalidSchema as err:
            print("Error de esquema inválido:", err.args[0])
            print("Comprueba que el protocolo es correcto. Ejemplo -> https://")
        except requests.exceptions.HTTPError as err:
            print("Informacion del error -> " + err.args[0])
        except requests.exceptions.ConnectionError as err:
            print(err.args[0])
            print("Error al conectarse. Intentos maximos superados.")
        except InvalidURL as err:
            print("URL inválida:", err.args[0])
            print("Comprueba la validez de la URL.")
        else:
            return r

    @staticmethod
    def push(menu_method):
        menu_stack.append(menu_method)

    @staticmethod
    def pop():
        if not Utils.isEmpty():
            menu_stack.pop()

    @staticmethod
    def isEmpty():
        return len(menu_stack) == 0
    
    # Method to separate the date fields
    # 2023-10-23T19:39:03Z
    @staticmethod
    def format_date(date):
        ymd = date.split("T")[0]
        time = date.split("T")[1].split("Z")[0]
        return str(ymd) + " " + str(time)

    #
    # Methods to download data and create different file formats
    #

    # Method to create a simple .txt with the retrieved data
    def create_txt(field_data_table, file_name, data, field_index, date_format):
        save_path = os.getcwd() + "/" + file_name + ".txt"
        
        try:
            if date_format == '1':
                pattern = r"│\s*(\d+)\s*│\s*(\d{4}-\d{2}-\d{2})\s*│\s*(\d{2}:\d{2}:\d{2})\s*│\s*(\d+\.\d+)\s*│"

                coincidencias = re.findall(pattern, field_data_table)

                all_rows = []
                for index, date, time, value in coincidencias:
                    row = []
                    row.append(index)
                    row.append(date + 'T' + time)
                    row.append(value)
                    all_rows.append(row)

                field_data_table = tabulate(all_rows, tablefmt='rounded_grid')

                with open(save_path, "w", encoding="utf-8") as file:
                    file.write(field_data_table)
                Utils.give_response(message="File created", status=True)  
        except Exception as e:
            Utils.give_response(message=f"File created {str(e)}", status=True)  


    # Method to create a simple .csv file with the field data
    def create_csv(field_data_table, file_name, data, field_index, date_format):
        store_path = os.getcwd() + "/" + file_name + ".csv"
        
        try:
            with open(store_path, "w") as file:
            # file.write("\t" + "   PRUEBA\n")
            # file.write("{:<12}{:<12}{}\n".format("Date", "Time", "Value"))
                for row in data:
                    if date_format == '1':
                        file.write(row['created_at'] + "\t" + row[f'field' + field_index] + "\n")
                    else:
                        date, t = Utils.format_date(row['created_at']).split(" ")
                        file.write(date + "\t" + t + "\t" + row[f'field' + field_index] + "\n")
            Utils.give_response(message="File created", status=True)        
        except Exception as e:
            Utils.give_response(message=f"File created {str(e)}", status=False)


    # Method to create a row in a excel sheet with given data
    def insert_row_in_sheet(ws, fila, datos):
        if type(datos) is list:
            # pdb.set_trace()
            col = 1
            for d in datos:
                ws.cell(row=fila, column=col, value=d)
                col += 1
        else:
            ws.cell(row=fila, column=1, value=datos)


    # Method to create a xlsx file for excel
    def create_xlsx(field_data_table, file_name, data, field_index, date_format):
        store_path = os.getcwd() + "/" + file_name + ".xlsx"

        try:
            wb = openpyxl.load_workbook(store_path)
            ws = wb.active
            ws.title = file_name

            # ws.merge_cells('A1:C1')
            # ws['A1'] = "PRUEBA"
            # Utils.introducir_fila_excel(ws, 1, "PRUEBA")
            # Utils.insert_row_in_sheet(ws, 2, ["Date", "Time", "Value"])

            row = 3
            for data_row in data:
                datetime = Utils.format_date(data_row['created_at'])
                date, time = datetime.split(" ")
                Utils.insert_row_in_sheet(ws, row, [date, time, data_row[f'field' + field_index]])
                row += 1

            wb.save(file_name + ".xlsx")
            Utils.give_response(message=f"File created", status=True)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            Utils.give_response(message=f"File created", status=False)
